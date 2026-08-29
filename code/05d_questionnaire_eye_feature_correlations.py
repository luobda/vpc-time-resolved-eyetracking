# -*- coding: utf-8 -*-
"""
六个标准化原始眼动指标与六个量表总分的 Spearman 相关分析。

分析流程：
1. 读取六个量表 Excel，提取每名被试的原始总分。
2. 读取 Table_S3_3_subject_six_features_standardized.csv。
3. 按手机号优先、唯一姓名其次的规则匹配被试。
4. 对每个“眼动指标 × 量表总分”组合执行 Spearman 秩相关。
5. 对全部相关检验执行 Benjamini-Hochberg FDR 校正。
6. 输出包含分析结果与匹配审计信息的 Excel 工作簿。

说明：
- 原始 Excel schema 保持中文；脚本不会修改源文件的数据结构。
- 六个量表均使用原始总分。
- 六个眼动指标使用上游已标准化的数值，不在本脚本中再次标准化。
- 注意控制、心力、MMSE：总分越高越好。
- 抑郁、焦虑、失眠：总分越低越好。
"""

from __future__ import annotations

import re
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from scipy import stats

warnings.filterwarnings("ignore", category=RuntimeWarning)


# =============================================================================
# 1. 路径设置
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()

SCALE_DIR = DATA_DIR

FEATURE_CSV = (
    SCRIPT_DIR
    / "Chapter3_Output"
    / "Table_S3_3_subject_six_features_standardized.csv"
)

OUTPUT_XLSX = (
    SCRIPT_DIR
    / "Chapter5_Output3"
    / "raw_eye_features_scale_spearman_results.xlsx"
)

# Spearman 相关最小有效样本量
MIN_CORRELATION_N = 10

# FDR 阈值
FDR_ALPHA = 0.05


# =============================================================================
# 2. 六个量表文件
# =============================================================================

SCALE_FILES = {
    "AttentionControl": "01_16条目注意控制量表_匿名化.xlsx",
    "MentalStrength": "02_心力自评量表_匿名化.xlsx",
    "Depression": "03_抑郁情绪自评量表_匿名化.xlsx",
    "Anxiety": "04_焦虑情绪自评量表_匿名化.xlsx",
    "Insomnia": "05_失眠严重程度指数自评量表_匿名化.xlsx",
    "MMSE": "06_简易精神状态评价量表_匿名化.xlsx",
}

SCALE_INFO = {
    "AttentionControl": {
        "scale_cn": "注意控制",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
    "MentalStrength": {
        "scale_cn": "心力",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
    "Depression": {
        "scale_cn": "抑郁",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "Anxiety": {
        "scale_cn": "焦虑",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "Insomnia": {
        "scale_cn": "失眠",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "MMSE": {
        "scale_cn": "简易精神状态 / MMSE",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
}

SCALE_COLUMNS = list(SCALE_FILES.keys())


# =============================================================================
# 3. 六个标准化原始眼动指标
# =============================================================================

FEATURE_COLUMNS = [
    "Memory Representation Ability",
    "Cognitive Processing Noise Level",
    "Memory Retrieval Latency",
    "Memory Representation Ability slope",
    "Cognitive Processing Noise slope",
    "Memory Retrieval Latency slope",
]

FEATURE_INFO = {
    "Memory Representation Ability": "平均表征水平",
    "Cognitive Processing Noise Level": "平均稳定性水平",
    "Memory Retrieval Latency": "平均提取水平",
    "Memory Representation Ability slope": "表征斜率",
    "Cognitive Processing Noise slope": "稳定性斜率",
    "Memory Retrieval Latency slope": "提取斜率",
}


# =============================================================================
# 4. 通用工具函数
# =============================================================================

def clean_column_name(value: Any) -> str:
    """清理列名。"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_name(value: Any) -> str:
    """标准化姓名。"""
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"[\s·•・]+", "", str(value).strip())


def normalize_phone(value: Any) -> str:
    """标准化手机号。"""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()

    if not text:
        return ""

    try:
        number = Decimal(text)
        if number == number.to_integral_value():
            text = str(int(number))
    except (InvalidOperation, ValueError):
        pass

    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D", "", text)

    if len(digits) > 11:
        digits = digits[-11:]

    return digits


def parse_subject(value: Any) -> tuple[str, str]:
    """
    从 subject 字段解析手机号和姓名。

    支持：
    - 张三
    - 138xxxx8888_张三
    - 138xxxx8888＿张三
    - subject 字段中混合手机号和姓名
    """
    text = "" if value is None or pd.isna(value) else str(value).strip()

    if not text:
        return "", ""

    parts = re.split(r"[_＿]", text, maxsplit=1)

    if len(parts) == 2:
        phone = normalize_phone(parts[0])
        name = normalize_name(parts[1])
        return phone, name

    phone_match = re.search(r"\d{7,}", text)
    phone = normalize_phone(phone_match.group(0)) if phone_match else ""

    name_text = re.sub(r"\d+", "", text)
    name = normalize_name(name_text)

    if not phone and not name:
        name = normalize_name(text)

    return phone, name


def first_existing_column(
    df: pd.DataFrame,
    candidates: Iterable[str],
) -> str | None:
    """寻找第一个存在的候选列名。"""
    lookup = {
        clean_column_name(column).lower(): column
        for column in df.columns
    }

    for candidate in candidates:
        key = clean_column_name(candidate).lower()
        if key in lookup:
            return lookup[key]

    return None


def read_csv_flexible(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV。"""
    last_error: Exception | None = None

    for encoding in ["utf-8-sig", "utf-8", "gb18030", "gbk"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"无法读取 CSV 文件：{path}") from last_error


def safe_numeric(series: pd.Series) -> pd.Series:
    """安全转换为数值。"""
    return pd.to_numeric(series, errors="coerce").replace(
        [np.inf, -np.inf],
        np.nan,
    )


def benjamini_hochberg(p_values: Iterable[float]) -> np.ndarray:
    """Benjamini-Hochberg FDR 校正。"""
    p = np.asarray(list(p_values), dtype=float)
    q = np.full_like(p, np.nan)

    valid = np.isfinite(p)

    if valid.sum() == 0:
        return q

    valid_p = p[valid]
    order = np.argsort(valid_p)
    ranked_p = valid_p[order]

    adjusted = ranked_p * len(ranked_p) / np.arange(1, len(ranked_p) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    adjusted = np.clip(adjusted, 0, 1)

    restored = np.empty_like(adjusted)
    restored[order] = adjusted
    q[valid] = restored

    return q


def ensure_inputs() -> None:
    """检查输入文件是否存在。"""
    required_files = [FEATURE_CSV]

    for filename in SCALE_FILES.values():
        required_files.append(SCALE_DIR / filename)

    missing = [str(path) for path in required_files if not path.exists()]

    if missing:
        raise FileNotFoundError(
            "以下输入文件不存在，请检查路径：\n"
            + "\n".join(missing)
        )

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)


# =============================================================================
# 5. 读取六个量表
# =============================================================================
# 原始 Excel 表头保持中文且不会被写回；内部字段仅用于内存中的分析与匹配。

def load_scale_workbook(
    scale_name: str,
    path: Path,
) -> pd.DataFrame:
    """
    读取单个量表文件，提取：
    - 姓名
    - 手机号
    - 总分
    - 填写日期

    如果同一人重复填写同一量表，则保留最新记录。
    """
    sheets = pd.read_excel(
        path,
        sheet_name=None,
        dtype=object,
        engine="openpyxl",
    )

    usable_tables: list[pd.DataFrame] = []

    for sheet_name, raw in sheets.items():
        raw = raw.copy()

        raw.columns = [
            clean_column_name(column)
            for column in raw.columns
        ]

        raw = (
            raw
            .dropna(axis=0, how="all")
            .dropna(axis=1, how="all")
        )

        name_column = first_existing_column(
            raw,
            ["姓名", "名字", "name", "Name"],
        )

        phone_column = first_existing_column(
            raw,
            ["手机号", "手机号码", "电话", "phone", "Phone"],
        )

        score_column = first_existing_column(
            raw,
            ["总分", "合计", "总得分", "score", "total", "Total"],
        )

        date_column = first_existing_column(
            raw,
            ["填写日期", "日期", "date", "Date"],
        )

        if name_column is None or score_column is None:
            continue

        part = pd.DataFrame({
            "scale": scale_name,
            "source_file": path.name,
            "source_sheet": sheet_name,
            "name_raw": raw[name_column],
            "phone_raw": raw[phone_column] if phone_column else "",
            "total_score": safe_numeric(raw[score_column]),
            "fill_date": raw[date_column] if date_column else np.nan,
        })

        part["name"] = part["name_raw"].map(normalize_name)
        part["phone"] = part["phone_raw"].map(normalize_phone)

        part = part[
            (part["name"] != "")
            | (part["phone"] != "")
        ].copy()

        usable_tables.append(part)

    if not usable_tables:
        raise ValueError(
            f"量表文件中没有找到可用的姓名列和总分列：{path}"
        )

    combined = pd.concat(usable_tables, ignore_index=True)

    combined["fill_date_parsed"] = pd.to_datetime(
        combined["fill_date"],
        errors="coerce",
    )

    combined["fill_date_numeric"] = pd.to_numeric(
        combined["fill_date"],
        errors="coerce",
    )

    combined["person_key"] = np.where(
        combined["phone"] != "",
        "P:" + combined["phone"],
        "N:" + combined["name"],
    )

    combined = combined.sort_values(
        [
            "person_key",
            "fill_date_parsed",
            "fill_date_numeric",
        ],
        na_position="first",
    )

    combined = combined.drop_duplicates(
        subset=["person_key"],
        keep="last",
    )

    combined = combined[
        combined["total_score"].notna()
    ].copy()

    return combined


def build_scale_master() -> pd.DataFrame:
    """读取六个量表并整理为宽格式。"""
    long_tables: list[pd.DataFrame] = []

    for scale_name, filename in SCALE_FILES.items():
        scale_path = SCALE_DIR / filename

        table = load_scale_workbook(
            scale_name=scale_name,
            path=scale_path,
        )

        long_tables.append(table)

    long_data = pd.concat(long_tables, ignore_index=True)

    score_wide = long_data.pivot_table(
        index="person_key",
        columns="scale",
        values="total_score",
        aggfunc="first",
    ).reset_index()

    identity = (
        long_data
        .sort_values(["person_key", "scale"])
        .groupby("person_key", as_index=False)
        .agg(
            phone=(
                "phone",
                lambda values: next((value for value in values if value), ""),
            ),
            name=(
                "name",
                lambda values: next((value for value in values if value), ""),
            ),
        )
    )

    master = identity.merge(
        score_wide,
        on="person_key",
        how="outer",
    )

    for scale in SCALE_COLUMNS:
        if scale not in master.columns:
            master[scale] = np.nan

        master[scale] = safe_numeric(master[scale])

    master["n_scale_available"] = master[SCALE_COLUMNS].notna().sum(axis=1)

    return master


# =============================================================================
# 6. 读取六个标准化原始眼动指标
# =============================================================================

def load_feature_file(path: Path) -> pd.DataFrame:
    """读取六个标准化原始眼动指标文件。"""
    data = read_csv_flexible(path)

    data = data.copy()
    data.columns = [
        clean_column_name(column)
        for column in data.columns
    ]

    subject_column = first_existing_column(
        data,
        [
            "subject",
            "Subject",
            "被试",
            "被试编号",
            "subject_id",
            "Subject_ID",
            "姓名",
            "name",
            "Name",
        ],
    )

    if subject_column is None:
        raise ValueError(
            "眼动指标文件中没有找到 subject / 被试 / 姓名 列。\n"
            "当前列名如下：\n"
            + "\n".join(map(str, data.columns))
        )

    data = data.rename(columns={subject_column: "subject"})

    missing_feature_columns = [
        column
        for column in FEATURE_COLUMNS
        if column not in data.columns
    ]

    if missing_feature_columns:
        raise ValueError(
            "眼动指标文件缺少以下列：\n"
            + "\n".join(missing_feature_columns)
            + "\n\n当前文件列名如下：\n"
            + "\n".join(map(str, data.columns))
        )

    parsed = data["subject"].apply(parse_subject)

    data["phone"] = parsed.map(lambda item: item[0])
    data["name"] = parsed.map(lambda item: item[1])

    # 如果原文件里本身有姓名或手机号列，也尝试补充使用。
    name_column = first_existing_column(
        data,
        ["姓名", "名字", "name", "Name"],
    )

    phone_column = first_existing_column(
        data,
        ["手机号", "手机号码", "电话", "phone", "Phone"],
    )

    if name_column is not None and name_column != "name":
        parsed_name = data[name_column].map(normalize_name)
        data["name"] = data["name"].where(data["name"] != "", parsed_name)

    if phone_column is not None and phone_column != "phone":
        parsed_phone = data[phone_column].map(normalize_phone)
        data["phone"] = data["phone"].where(data["phone"] != "", parsed_phone)

    for column in FEATURE_COLUMNS:
        data[column] = safe_numeric(data[column])

    data = data[
        (data["name"] != "")
        | (data["phone"] != "")
    ].copy()

    data = data.drop_duplicates(
        subset=["subject"],
        keep="last",
    )

    return data


# =============================================================================
# 7. 匹配量表数据与眼动指标数据
# =============================================================================

def match_scale_to_features(
    scale_master: pd.DataFrame,
    feature_data: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    将量表数据与眼动指标数据匹配。

    匹配优先级：
    1. 手机号唯一匹配
    2. 姓名唯一匹配
    """
    scale = scale_master.copy()
    features = feature_data.copy()

    unique_phone_to_key = (
        scale[scale["phone"] != ""]
        .drop_duplicates("phone", keep=False)
        .set_index("phone")["person_key"]
        .to_dict()
    )

    unique_name_to_key = (
        scale[scale["name"] != ""]
        .drop_duplicates("name", keep=False)
        .set_index("name")["person_key"]
        .to_dict()
    )

    def resolve(row: pd.Series) -> tuple[str | None, str]:
        phone = row.get("phone", "")
        name = row.get("name", "")

        match_by_phone = unique_phone_to_key.get(phone) if phone else None
        match_by_name = unique_name_to_key.get(name) if name else None

        if (
            match_by_phone is not None
            and match_by_name is not None
            and match_by_phone == match_by_name
        ):
            return match_by_phone, "phone+name"

        if match_by_phone is not None:
            return match_by_phone, "phone"

        if match_by_name is not None:
            return match_by_name, "unique_name"

        return None, "unmatched"

    resolved = features.apply(resolve, axis=1)

    features["person_key"] = resolved.map(lambda item: item[0])
    features["match_method"] = resolved.map(lambda item: item[1])

    merged = features.merge(
        scale,
        on="person_key",
        how="left",
        suffixes=("_feature", "_scale"),
    )

    audit_columns = [
        column
        for column in [
            "subject",
            "phone_feature",
            "name_feature",
            "phone_scale",
            "name_scale",
            "person_key",
            "match_method",
        ]
        if column in merged.columns
    ]

    match_audit = merged[audit_columns].copy()
    match_audit["matched"] = merged["person_key"].notna()

    return merged, match_audit


# =============================================================================
# 8. Spearman 相关分析
# =============================================================================

def interpret_spearman_result(
    feature: str,
    scale: str,
    rho: float,
) -> str:
    """
    根据量表方向解释 Spearman rho。

    注意：
    这里解释的是相关方向，不代表因果关系。
    """
    if not np.isfinite(rho):
        return ""

    feature_cn = FEATURE_INFO.get(feature, feature)
    scale_cn = SCALE_INFO[scale]["scale_cn"]
    direction = SCALE_INFO[scale]["score_direction"]

    if abs(rho) < 1e-12:
        return f"{feature_cn}与{scale_cn}总分几乎无单调相关"

    if direction == "higher_is_better":
        if rho > 0:
            return f"{feature_cn}越高，{scale_cn}总分越高，通常表示状态越好"
        else:
            return f"{feature_cn}越高，{scale_cn}总分越低，通常表示状态越差"

    if direction == "lower_is_better":
        if rho > 0:
            return f"{feature_cn}越高，{scale_cn}总分越高，通常表示症状更重 / 状态更差"
        else:
            return f"{feature_cn}越高，{scale_cn}总分越低，通常表示症状更轻 / 状态更好"

    return ""


def run_spearman_feature_analysis(
    data: pd.DataFrame,
    feature_columns: list[str],
    scale_columns: list[str],
) -> pd.DataFrame:
    """对每个眼动指标与每个量表总分做 Spearman 相关。"""
    rows: list[dict[str, Any]] = []

    for feature in feature_columns:
        for scale in scale_columns:
            if feature not in data.columns or scale not in data.columns:
                continue

            pair = pd.DataFrame({
                "feature_value": safe_numeric(data[feature]),
                "scale_score": safe_numeric(data[scale]),
            }).dropna()

            n = len(pair)
            rho = np.nan
            p_value = np.nan

            if (
                n >= MIN_CORRELATION_N
                and pair["feature_value"].nunique() >= 2
                and pair["scale_score"].nunique() >= 2
            ):
                rho, p_value = stats.spearmanr(
                    pair["feature_value"],
                    pair["scale_score"],
                )

            rows.append({
                "predictor": feature,
                "predictor_cn": FEATURE_INFO.get(feature, feature),
                "scale": scale,
                "scale_cn": SCALE_INFO[scale]["scale_cn"],
                "score_direction": SCALE_INFO[scale]["score_direction"],
                "direction_cn": SCALE_INFO[scale]["direction_cn"],
                "n": n,
                "spearman_rho": float(rho) if np.isfinite(rho) else np.nan,
                "p": float(p_value) if np.isfinite(p_value) else np.nan,
                "interpretation": interpret_spearman_result(feature, scale, rho),
            })

    results = pd.DataFrame(rows)

    if not results.empty:
        results["q_FDR"] = benjamini_hochberg(results["p"])
        results["FDR_significant"] = results["q_FDR"] < FDR_ALPHA
        results["nominal_p_lt_05"] = results["p"] < 0.05

        results = results.sort_values(
            ["q_FDR", "p", "predictor", "scale"],
            na_position="last",
            ignore_index=True,
        )

    return results


# =============================================================================
# 9. Excel 美化
# =============================================================================

def style_excel_workbook(path: Path) -> None:
    """简单美化 Excel 输出。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            preview = list(column_cells)[:200]

            max_length = max(
                (
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                    for cell in preview
                ),
                default=0,
            )

            width = min(max(max_length + 2, 10), 42)
            worksheet.column_dimensions[preview[0].column_letter].width = width

    workbook.save(path)

# =============================================================================
# 10. 汇总与导出
# =============================================================================

def build_output_tables(
    scale_master: pd.DataFrame,
    feature_data: pd.DataFrame,
    analysis_data: pd.DataFrame,
    match_audit: pd.DataFrame,
    spearman_results: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """构建最终 Excel 工作簿中的全部结果表。"""
    nominal_results = spearman_results[
        spearman_results["nominal_p_lt_05"] == True
    ].copy()

    fdr_results = spearman_results[
        spearman_results["FDR_significant"] == True
    ].copy()

    count_table = pd.DataFrame([
        {
            "item": "scale_participants",
            "value": len(scale_master),
        },
        {
            "item": "feature_participants",
            "value": len(feature_data),
        },
        {
            "item": "matched_participants",
            "value": int(match_audit["matched"].sum()),
        },
        {
            "item": "participants_with_all_6_scales_after_matching",
            "value": int(
                analysis_data[SCALE_COLUMNS]
                .notna()
                .all(axis=1)
                .sum()
            ),
        },
        {
            "item": "feature_columns_count",
            "value": len(FEATURE_COLUMNS),
        },
        {
            "item": "spearman_tests_count",
            "value": len(spearman_results),
        },
        {
            "item": "nominal_p_lt_05_count",
            "value": len(nominal_results),
        },
        {
            "item": "fdr_q_lt_05_count",
            "value": len(fdr_results),
        },
    ])

    scale_rule_table = pd.DataFrame([
        {
            "scale": scale,
            "scale_cn": SCALE_INFO[scale]["scale_cn"],
            "data_used": "raw_total_score",
            "score_direction": SCALE_INFO[scale]["score_direction"],
            "direction_cn": SCALE_INFO[scale]["direction_cn"],
        }
        for scale in SCALE_COLUMNS
    ])

    feature_rule_table = pd.DataFrame([
        {
            "predictor": feature,
            "predictor_cn": FEATURE_INFO.get(feature, feature),
            "data_used": "standardized_eye_feature",
        }
        for feature in FEATURE_COLUMNS
    ])

    readme_table = pd.DataFrame({
        "item": [
            "analysis",
            "predictor",
            "outcome",
            "method",
            "feature_columns_used",
            "feature_transformation",
            "scale_transformation",
            "multiple_comparison",
            "minimum_n",
            "interpretation_note",
        ],
        "value": [
            "Six standardized raw eye movement features vs questionnaire total score",
            "Six standardized raw eye movement features",
            "Six questionnaire raw total scores",
            "Spearman rank correlation",
            ", ".join(FEATURE_COLUMNS),
            "Features are already standardized; used directly",
            "No PCA, no z-score, no reverse transformation for scale scores",
            "Benjamini-Hochberg FDR across all feature × scale tests",
            MIN_CORRELATION_N,
            (
                "Depression/Anxiety/Insomnia: lower is better; "
                "AttentionControl/MentalStrength/MMSE: higher is better; "
                "correlation does not imply causality"
            ),
        ],
    })

    matched_columns = [
        column
        for column in [
            "subject",
            "person_key",
            "match_method",
            *FEATURE_COLUMNS,
            *SCALE_COLUMNS,
        ]
        if column in analysis_data.columns
    ]
    matched_data_compact = analysis_data[matched_columns].copy()

    return {
        "00_README": readme_table,
        "01_scale_rules": scale_rule_table,
        "02_feature_rules": feature_rule_table,
        "03_sample_counts": count_table,
        "04_matched_data": matched_data_compact,
        "05_spearman_results": spearman_results,
        "06_nominal_p_lt_05": nominal_results,
        "07_fdr_significant": fdr_results,
        "08_match_audit": match_audit,
    }


def export_tables(
    output_path: Path,
    tables: dict[str, pd.DataFrame],
) -> None:
    """将结果表写入一个 Excel 工作簿并统一设置格式。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, table in tables.items():
            safe_table = table.replace([np.inf, -np.inf], np.nan)
            safe_table.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    style_excel_workbook(output_path)


# =============================================================================
# 11. 主程序
# =============================================================================

def main() -> None:
    """运行完整的量表总分 × 标准化原始眼动指标相关分析。"""
    ensure_inputs()

    # 1. 读取六个量表总分
    scale_master = build_scale_master()

    # 2. 读取六个标准化原始眼动指标
    feature_data = load_feature_file(FEATURE_CSV)

    print("本次分析使用的六个标准化原始眼动指标：")
    for column in FEATURE_COLUMNS:
        print(f"  - {column}")

    # 3. 匹配量表与眼动指标数据
    merged, match_audit = match_scale_to_features(
        scale_master=scale_master,
        feature_data=feature_data,
    )

    matched_mask = merged["person_key"].notna()
    complete_scale_mask = merged[SCALE_COLUMNS].notna().all(axis=1)
    analysis_mask = matched_mask & complete_scale_mask

    analysis_data = merged[analysis_mask].copy()

    # 4. Spearman 相关分析
    spearman_results = run_spearman_feature_analysis(
        data=analysis_data,
        feature_columns=FEATURE_COLUMNS,
        scale_columns=SCALE_COLUMNS,
    )

    # 5. 汇总并导出
    tables = build_output_tables(
        scale_master=scale_master,
        feature_data=feature_data,
        analysis_data=analysis_data,
        match_audit=match_audit,
        spearman_results=spearman_results,
    )
    export_tables(OUTPUT_XLSX, tables)

    nominal_n = len(tables["06_nominal_p_lt_05"])
    fdr_n = len(tables["07_fdr_significant"])

    print("\n分析完成。")
    print(f"输出文件：{OUTPUT_XLSX}")
    print(f"成功匹配被试数：{int(match_audit['matched'].sum())}")
    print(f"Spearman 检验数量：{len(spearman_results)}")
    print(f"未校正 p < .05 的结果数量：{nominal_n}")
    print(f"FDR q < .05 的结果数量：{fdr_n}")


if __name__ == "__main__":
    main()
