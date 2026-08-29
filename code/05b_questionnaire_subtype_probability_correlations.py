# -*- coding: utf-8 -*-
"""
人群亚型概率与六个量表总分的 Spearman 相关分析。

分析流程：
1. 读取六个量表 Excel，提取每名被试的原始总分。
2. 读取人群亚型概率 CSV。
3. 按手机号优先、唯一姓名其次的规则匹配被试。
4. 仅保留完成全部六个量表且成功匹配的被试。
5. 对每个“亚型概率 × 量表总分”组合执行 Spearman 秩相关。
6. 对全部相关检验执行 Benjamini-Hochberg FDR 校正。
7. 输出一个包含分析结果与匹配审计信息的 Excel 工作簿。

说明：
- 原始 Excel schema 保持不变；脚本只在内存中建立分析字段。
- 六个量表均使用原始总分。
- 不做 PCA、不做 z 标准化、不做反向转换。
- 注意控制、心力、MMSE：总分越高越好。
- 抑郁、焦虑、失眠：总分越低越好。
"""

from __future__ import annotations

import re
import warnings
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from scipy import stats

warnings.filterwarnings("ignore", category=RuntimeWarning)


# =============================================================================
# 1. 路径设置
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()

SCALE_DIR = DATA_DIR

PROBABILITY_CSV = (
    SCRIPT_DIR
    / "Chapter4_Output"
    / "Table_S4_subject_group_membership_probabilities.csv"
)

OUTPUT_XLSX = (
    SCRIPT_DIR
    / "Chapter5_Output1"
    / "subtype_probability_scale_spearman_results1.xlsx"
)

# 如果你的概率列名比较特殊，自动识别失败，可以在这里手动指定。
# 例如：
MANUAL_PROBABILITY_COLUMNS = [
    "prob_low_representation",
    "prob_representation_maintenance",
    "prob_stability_retrieval_compensation",
    "prob_retrieval_inefficient_stability_compensation",
]


# Spearman 相关最小有效样本量
MIN_CORRELATION_N = 10

# FDR 阈值
FDR_ALPHA = 0.05


# =============================================================================
# 2. 六个量表文件与方向
# =============================================================================

SCALE_FILES = {
    "AttentionControl": "01_16条目注意控制量表_匿名化.xlsx",
    "MentalStrength": "02_心力自评量表_匿名化.xlsx",
    "Depression": "03_抑郁情绪自评量表_匿名化.xlsx",
    "Anxiety": "04_焦虑情绪自评量表_匿名化.xlsx",
    "Insomnia": "05_失眠严重程度指数自评量表_匿名化.xlsx",
    "MMSE": "06_简易精神状态评价量表_匿名化.xlsx",
}

SCALE_INFO = {
    "AttentionControl": {
        "scale_cn": "注意控制",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
    "MentalStrength": {
        "scale_cn": "心力",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
    "Depression": {
        "scale_cn": "抑郁",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "Anxiety": {
        "scale_cn": "焦虑",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "Insomnia": {
        "scale_cn": "失眠",
        "score_direction": "lower_is_better",
        "direction_cn": "总分越低越好",
    },
    "MMSE": {
        "scale_cn": "简易精神状态 / MMSE",
        "score_direction": "higher_is_better",
        "direction_cn": "总分越高越好",
    },
}

SCALE_COLUMNS = list(SCALE_FILES.keys())


# =============================================================================
# 3. 通用工具函数
# =============================================================================

def clean_column_name(value: Any) -> str:
    """清理列名。"""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_name(value: Any) -> str:
    """标准化姓名：去掉空格、特殊分隔符。"""
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"[\s·•・]+", "", str(value).strip())


def normalize_phone(value: Any) -> str:
    """标准化手机号。"""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    try:
        number = Decimal(text)
        if number == number.to_integral_value():
            text = str(int(number))
    except (InvalidOperation, ValueError):
        pass

    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D", "", text)

    # 如果前面带了国家码或其他字符，只保留最后 11 位。
    if len(digits) > 11:
        digits = digits[-11:]

    return digits


def parse_subject(value: Any) -> tuple[str, str]:
    """
    从 subject 字段里解析手机号和姓名。

    支持：
    - 138xxxx8888_张三
    - 138xxxx8888＿张三
    - 张三
    - subject 字段里混有手机号和姓名
    """
    text = "" if value is None or pd.isna(value) else str(value).strip()
    if not text:
        return "", ""

    parts = re.split(r"[_＿]", text, maxsplit=1)

    if len(parts) == 2:
        phone = normalize_phone(parts[0])
        name = normalize_name(parts[1])
        return phone, name

    phone_match = re.search(r"\d{7,}", text)
    phone = normalize_phone(phone_match.group(0)) if phone_match else ""

    name_text = re.sub(r"\d+", "", text)
    name = normalize_name(name_text)

    # 如果整个字段看起来就是姓名
    if not phone and not name:
        name = normalize_name(text)

    return phone, name


def first_existing_column(
    df: pd.DataFrame,
    candidates: Iterable[str],
) -> str | None:
    """在 DataFrame 中寻找第一个存在的候选列名。"""
    lookup = {
        clean_column_name(column).lower(): column
        for column in df.columns
    }

    for candidate in candidates:
        key = clean_column_name(candidate).lower()
        if key in lookup:
            return lookup[key]

    return None


def read_csv_flexible(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV。"""
    last_error: Exception | None = None

    for encoding in ["utf-8-sig", "utf-8", "gb18030", "gbk"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"无法读取 CSV 文件：{path}") from last_error


def safe_numeric(series: pd.Series) -> pd.Series:
    """转换为数值，非法值转为 NaN。"""
    return pd.to_numeric(series, errors="coerce").replace(
        [np.inf, -np.inf],
        np.nan,
    )


def benjamini_hochberg(p_values: Iterable[float]) -> np.ndarray:
    """Benjamini-Hochberg FDR 校正。"""
    p = np.asarray(list(p_values), dtype=float)
    q = np.full_like(p, np.nan)

    valid = np.isfinite(p)

    if valid.sum() == 0:
        return q

    valid_p = p[valid]
    order = np.argsort(valid_p)
    ranked_p = valid_p[order]

    adjusted = ranked_p * len(ranked_p) / np.arange(1, len(ranked_p) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    adjusted = np.clip(adjusted, 0, 1)

    restored = np.empty_like(adjusted)
    restored[order] = adjusted
    q[valid] = restored

    return q


def ensure_inputs() -> None:
    """检查输入文件是否存在。"""
    required_files = [PROBABILITY_CSV]

    for filename in SCALE_FILES.values():
        required_files.append(SCALE_DIR / filename)

    missing = [str(path) for path in required_files if not path.exists()]

    if missing:
        raise FileNotFoundError(
            "以下输入文件不存在，请检查路径：\n"
            + "\n".join(missing)
        )

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)


# =============================================================================
# 4. 读取六个量表
# =============================================================================
# 原始 Excel 表头保持中文且不会被写回或修改；以下处理仅发生在内存 DataFrame 中。

def load_scale_workbook(
    scale_name: str,
    path: Path,
) -> pd.DataFrame:
    """
    读取单个量表文件，提取：
    - 姓名
    - 手机号
    - 总分
    - 填写日期

    如果同一个人重复填写同一个量表，则保留最新记录。
    """
    sheets = pd.read_excel(
        path,
        sheet_name=None,
        dtype=object,
        engine="openpyxl",
    )

    usable_tables: list[pd.DataFrame] = []

    for sheet_name, raw in sheets.items():
        raw = raw.copy()

        raw.columns = [
            clean_column_name(column)
            for column in raw.columns
        ]

        raw = (
            raw
            .dropna(axis=0, how="all")
            .dropna(axis=1, how="all")
        )

        name_column = first_existing_column(
            raw,
            ["姓名", "名字", "name", "Name"],
        )

        phone_column = first_existing_column(
            raw,
            ["手机号", "手机号码", "电话", "phone", "Phone"],
        )

        score_column = first_existing_column(
            raw,
            ["总分", "合计", "总得分", "score", "total", "Total"],
        )

        date_column = first_existing_column(
            raw,
            ["填写日期", "日期", "date", "Date"],
        )

        if name_column is None or score_column is None:
            continue

        part = pd.DataFrame({
            "scale": scale_name,
            "source_file": path.name,
            "source_sheet": sheet_name,
            "name_raw": raw[name_column],
            "phone_raw": raw[phone_column] if phone_column else "",
            "total_score": safe_numeric(raw[score_column]),
            "fill_date": raw[date_column] if date_column else np.nan,
        })

        part["name"] = part["name_raw"].map(normalize_name)
        part["phone"] = part["phone_raw"].map(normalize_phone)

        part = part[
            (part["name"] != "")
            | (part["phone"] != "")
        ].copy()

        usable_tables.append(part)

    if not usable_tables:
        raise ValueError(
            f"量表文件中没有找到可用的姓名列和总分列：{path}"
        )

    combined = pd.concat(usable_tables, ignore_index=True)

    combined["fill_date_parsed"] = pd.to_datetime(
        combined["fill_date"],
        errors="coerce",
    )

    combined["fill_date_numeric"] = pd.to_numeric(
        combined["fill_date"],
        errors="coerce",
    )

    combined["person_key"] = np.where(
        combined["phone"] != "",
        "P:" + combined["phone"],
        "N:" + combined["name"],
    )

    combined = combined.sort_values(
        [
            "person_key",
            "fill_date_parsed",
            "fill_date_numeric",
        ],
        na_position="first",
    )

    # 同一人同一量表多次填写时，保留最新一次。
    combined = combined.drop_duplicates(
        subset=["person_key"],
        keep="last",
    )

    combined = combined[
        combined["total_score"].notna()
    ].copy()

    return combined


def build_scale_master() -> pd.DataFrame:
    """读取六个量表并整理为宽格式。"""
    long_tables: list[pd.DataFrame] = []

    for scale_name, filename in SCALE_FILES.items():
        scale_path = SCALE_DIR / filename

        table = load_scale_workbook(
            scale_name=scale_name,
            path=scale_path,
        )

        long_tables.append(table)

    long_data = pd.concat(long_tables, ignore_index=True)

    score_wide = long_data.pivot_table(
        index="person_key",
        columns="scale",
        values="total_score",
        aggfunc="first",
    ).reset_index()

    identity = (
        long_data
        .sort_values(["person_key", "scale"])
        .groupby("person_key", as_index=False)
        .agg(
            phone=(
                "phone",
                lambda values: next((value for value in values if value), ""),
            ),
            name=(
                "name",
                lambda values: next((value for value in values if value), ""),
            ),
        )
    )

    master = identity.merge(
        score_wide,
        on="person_key",
        how="outer",
    )

    for scale in SCALE_COLUMNS:
        if scale not in master.columns:
            master[scale] = np.nan

        master[scale] = safe_numeric(master[scale])

    master["n_scale_available"] = master[SCALE_COLUMNS].notna().sum(axis=1)

    return master


# =============================================================================
# 5. 读取人群亚型概率文件
# =============================================================================

def detect_probability_columns(df: pd.DataFrame) -> list[str]:
    """
    自动识别亚型概率列。

    优先识别列名中包含：
    - prob
    - probability
    - 概率
    - membership

    如果没有识别到，再尝试从数值范围 0~1 的列中推断。
    """
    if MANUAL_PROBABILITY_COLUMNS is not None:
        missing = [
            column
            for column in MANUAL_PROBABILITY_COLUMNS
            if column not in df.columns
        ]

        if missing:
            raise ValueError(
                "MANUAL_PROBABILITY_COLUMNS 中以下列不存在：\n"
                + "\n".join(missing)
            )

        return MANUAL_PROBABILITY_COLUMNS

    probability_patterns = [
        r"prob",
        r"probability",
        r"概率",
        r"membership",
    ]

    probability_columns: list[str] = []

    for column in df.columns:
        column_text = str(column).lower()

        if any(re.search(pattern, column_text) for pattern in probability_patterns):
            probability_columns.append(column)

    if probability_columns:
        return probability_columns

    # 兜底方案：尝试寻找数值范围基本在 0~1 的列。
    excluded_keywords = [
        "subject",
        "name",
        "姓名",
        "phone",
        "手机号",
        "cluster",
        "subtype",
        "label",
        "类别",
        "亚型",
        "group",
        "组别",
        "screen",
    ]

    numeric_probability_like: list[str] = []

    for column in df.columns:
        column_text = str(column).lower()

        if any(keyword.lower() in column_text for keyword in excluded_keywords):
            continue

        values = safe_numeric(df[column]).dropna()

        if len(values) == 0:
            continue

        if values.between(0, 1).mean() >= 0.95:
            numeric_probability_like.append(column)

    if len(numeric_probability_like) >= 2:
        return numeric_probability_like

    raise ValueError(
        "未能自动识别人群亚型概率列。\n"
        "请在代码顶部手动设置 MANUAL_PROBABILITY_COLUMNS。\n"
        "当前文件列名如下：\n"
        + "\n".join(map(str, df.columns))
    )


def load_probability_file(path: Path) -> tuple[pd.DataFrame, list[str]]:
    """读取人群亚型概率文件。"""
    data = read_csv_flexible(path)

    data = data.copy()
    data.columns = [
        clean_column_name(column)
        for column in data.columns
    ]

    subject_column = first_existing_column(
        data,
        [
            "subject",
            "Subject",
            "被试",
            "被试编号",
            "subject_id",
            "Subject_ID",
            "姓名",
            "name",
            "Name",
        ],
    )

    if subject_column is None:
        raise ValueError(
            "概率文件中没有找到 subject / 被试 / 姓名 列。\n"
            "当前列名如下：\n"
            + "\n".join(map(str, data.columns))
        )

    data = data.rename(columns={subject_column: "subject"})

    parsed = data["subject"].apply(parse_subject)

    data["phone"] = parsed.map(lambda item: item[0])
    data["name"] = parsed.map(lambda item: item[1])

    # 如果原文件里本身有姓名或手机号列，也一并尝试读取。
    name_column = first_existing_column(
        data,
        ["姓名", "名字", "name", "Name"],
    )

    phone_column = first_existing_column(
        data,
        ["手机号", "手机号码", "电话", "phone", "Phone"],
    )

    if name_column is not None and name_column != "name":
        parsed_name = data[name_column].map(normalize_name)
        data["name"] = data["name"].where(data["name"] != "", parsed_name)

    if phone_column is not None and phone_column != "phone":
        parsed_phone = data[phone_column].map(normalize_phone)
        data["phone"] = data["phone"].where(data["phone"] != "", parsed_phone)

    probability_columns = detect_probability_columns(data)

    for column in probability_columns:
        data[column] = safe_numeric(data[column])

    # 去掉完全没有姓名和手机号的记录。
    data = data[
        (data["name"] != "")
        | (data["phone"] != "")
    ].copy()

    # 同一 subject 重复时保留最后一条。
    data = data.drop_duplicates(
        subset=["subject"],
        keep="last",
    )

    return data, probability_columns


# =============================================================================
# 6. 匹配量表数据与概率数据
# =============================================================================

def match_scale_to_probability(
    scale_master: pd.DataFrame,
    probability_data: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    将量表数据与概率数据匹配。

    优先级：
    1. 手机号唯一匹配
    2. 姓名唯一匹配
    """
    scale = scale_master.copy()
    prob = probability_data.copy()

    unique_phone_to_key = (
        scale[scale["phone"] != ""]
        .drop_duplicates("phone", keep=False)
        .set_index("phone")["person_key"]
        .to_dict()
    )

    unique_name_to_key = (
        scale[scale["name"] != ""]
        .drop_duplicates("name", keep=False)
        .set_index("name")["person_key"]
        .to_dict()
    )

    def resolve(row: pd.Series) -> tuple[str | None, str]:
        phone = row.get("phone", "")
        name = row.get("name", "")

        match_by_phone = unique_phone_to_key.get(phone) if phone else None
        match_by_name = unique_name_to_key.get(name) if name else None

        if (
            match_by_phone is not None
            and match_by_name is not None
            and match_by_phone == match_by_name
        ):
            return match_by_phone, "phone+name"

        if match_by_phone is not None:
            return match_by_phone, "phone"

        if match_by_name is not None:
            return match_by_name, "unique_name"

        return None, "unmatched"

    resolved = prob.apply(resolve, axis=1)

    prob["person_key"] = resolved.map(lambda item: item[0])
    prob["match_method"] = resolved.map(lambda item: item[1])

    merged = prob.merge(
        scale,
        on="person_key",
        how="left",
        suffixes=("_probability", "_scale"),
    )

    audit_columns = [
        column
        for column in [
            "subject",
            "phone_probability",
            "name_probability",
            "phone_scale",
            "name_scale",
            "person_key",
            "match_method",
        ]
        if column in merged.columns
    ]

    match_audit = merged[audit_columns].copy()
    match_audit["matched"] = merged["person_key"].notna()

    return merged, match_audit


# =============================================================================
# 7. Spearman 相关分析
# =============================================================================

def interpret_spearman_result(
    scale: str,
    rho: float,
) -> str:
    """
    根据量表方向解释 Spearman rho 的方向。

    注意：
    这里解释的是相关方向，不代表因果关系。
    """
    if not np.isfinite(rho):
        return ""

    scale_cn = SCALE_INFO[scale]["scale_cn"]
    direction = SCALE_INFO[scale]["score_direction"]

    if abs(rho) < 1e-12:
        return f"亚型概率与{scale_cn}总分几乎无单调相关"

    if direction == "higher_is_better":
        if rho > 0:
            return f"亚型概率越高，{scale_cn}总分越高，通常表示状态越好"
        else:
            return f"亚型概率越高，{scale_cn}总分越低，通常表示状态越差"

    if direction == "lower_is_better":
        if rho > 0:
            return f"亚型概率越高，{scale_cn}总分越高，通常表示症状更重 / 状态更差"
        else:
            return f"亚型概率越高，{scale_cn}总分越低，通常表示症状更轻 / 状态更好"

    return ""


def run_spearman_probability_analysis(
    data: pd.DataFrame,
    probability_columns: list[str],
    scale_columns: list[str],
) -> pd.DataFrame:
    """对每个亚型概率与每个量表总分做 Spearman 相关。"""
    rows: list[dict[str, Any]] = []

    for probability_column in probability_columns:
        for scale in scale_columns:
            if probability_column not in data.columns or scale not in data.columns:
                continue

            pair = pd.DataFrame({
                "probability": safe_numeric(data[probability_column]),
                "scale_score": safe_numeric(data[scale]),
            }).dropna()

            n = len(pair)
            rho = np.nan
            p_value = np.nan

            if (
                n >= MIN_CORRELATION_N
                and pair["probability"].nunique() >= 2
                and pair["scale_score"].nunique() >= 2
            ):
                rho, p_value = stats.spearmanr(
                    pair["probability"],
                    pair["scale_score"],
                )

            rows.append({
                "predictor": probability_column,
                "scale": scale,
                "scale_cn": SCALE_INFO[scale]["scale_cn"],
                "score_direction": SCALE_INFO[scale]["score_direction"],
                "direction_cn": SCALE_INFO[scale]["direction_cn"],
                "n": n,
                "spearman_rho": float(rho) if np.isfinite(rho) else np.nan,
                "p": float(p_value) if np.isfinite(p_value) else np.nan,
                "interpretation": interpret_spearman_result(scale, rho),
            })

    results = pd.DataFrame(rows)

    if not results.empty:
        results["q_FDR"] = benjamini_hochberg(results["p"])
        results["FDR_significant"] = results["q_FDR"] < FDR_ALPHA
        results["nominal_p_lt_05"] = results["p"] < 0.05

        results = results.sort_values(
            ["q_FDR", "p", "predictor", "scale"],
            na_position="last",
            ignore_index=True,
        )

    return results


# =============================================================================
# 8. Excel 美化
# =============================================================================

def style_excel_workbook(path: Path) -> None:
    """简单美化 Excel 输出。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in worksheet.columns:
            preview = list(column_cells)[:200]
            max_length = max(
                (
                    len(str(cell.value))
                    if cell.value is not None
                    else 0
                    for cell in preview
                ),
                default=0,
            )

            width = min(max(max_length + 2, 10), 40)
            worksheet.column_dimensions[preview[0].column_letter].width = width

    workbook.save(path)

# =============================================================================
# 9. 汇总与导出
# =============================================================================

def build_output_tables(
    scale_master: pd.DataFrame,
    probability_data: pd.DataFrame,
    analysis_data: pd.DataFrame,
    match_audit: pd.DataFrame,
    probability_columns: list[str],
    spearman_results: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """构建最终 Excel 工作簿中的全部表格。"""
    nominal_results = spearman_results[
        spearman_results["nominal_p_lt_05"] == True
    ].copy()

    fdr_results = spearman_results[
        spearman_results["FDR_significant"] == True
    ].copy()

    count_table = pd.DataFrame([
        {
            "item": "scale_participants",
            "value": len(scale_master),
        },
        {
            "item": "probability_participants",
            "value": len(probability_data),
        },
        {
            "item": "matched_participants",
            "value": int(match_audit["matched"].sum()),
        },
        {
            "item": "participants_with_all_6_scales_after_matching",
            "value": int(
                analysis_data[SCALE_COLUMNS]
                .notna()
                .all(axis=1)
                .sum()
            ),
        },
        {
            "item": "probability_columns_count",
            "value": len(probability_columns),
        },
        {
            "item": "spearman_tests_count",
            "value": len(spearman_results),
        },
        {
            "item": "nominal_p_lt_05_count",
            "value": len(nominal_results),
        },
        {
            "item": "fdr_q_lt_05_count",
            "value": len(fdr_results),
        },
    ])

    scale_rule_table = pd.DataFrame([
        {
            "scale": scale,
            "scale_cn": SCALE_INFO[scale]["scale_cn"],
            "data_used": "raw_total_score",
            "score_direction": SCALE_INFO[scale]["score_direction"],
            "direction_cn": SCALE_INFO[scale]["direction_cn"],
        }
        for scale in SCALE_COLUMNS
    ])

    readme_table = pd.DataFrame({
        "item": [
            "analysis",
            "predictor",
            "outcome",
            "method",
            "scale_transformation",
            "multiple_comparison",
            "minimum_n",
            "interpretation_note",
        ],
        "value": [
            "Subtype membership probability vs questionnaire total score",
            "Probability of belonging to each subtype",
            "Six questionnaire raw total scores",
            "Spearman rank correlation",
            "No PCA, no z-score, no reverse transformation",
            "Benjamini-Hochberg FDR across all probability × scale tests",
            MIN_CORRELATION_N,
            (
                "Depression/Anxiety/Insomnia: lower is better; "
                "AttentionControl/MentalStrength/MMSE: higher is better; "
                "correlation does not imply causality"
            ),
        ],
    })

    matched_columns = [
        column
        for column in [
            "subject",
            "person_key",
            "match_method",
            *probability_columns,
            *SCALE_COLUMNS,
        ]
        if column in analysis_data.columns
    ]
    matched_data_compact = analysis_data[matched_columns].copy()

    return {
        "00_README": readme_table,
        "01_scale_rules": scale_rule_table,
        "02_sample_counts": count_table,
        "03_matched_data": matched_data_compact,
        "04_spearman_results": spearman_results,
        "05_nominal_p_lt_05": nominal_results,
        "06_fdr_significant": fdr_results,
        "07_match_audit": match_audit,
    }


def export_tables(
    output_path: Path,
    tables: dict[str, pd.DataFrame],
) -> None:
    """将全部结果表写入同一个 Excel 工作簿并统一设置格式。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, table in tables.items():
            safe_table = table.replace([np.inf, -np.inf], np.nan)
            safe_table.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

    style_excel_workbook(output_path)


# =============================================================================
# 10. 主程序
# =============================================================================

def main() -> None:
    """运行完整的量表总分 × 人群亚型概率相关分析。"""
    ensure_inputs()

    # 1. 读取六个量表总分
    scale_master = build_scale_master()

    # 2. 读取人群亚型概率
    probability_data, probability_columns = load_probability_file(PROBABILITY_CSV)

    print("识别到的亚型概率列：")
    for column in probability_columns:
        print(f"  - {column}")

    # 3. 匹配量表与概率数据
    merged, match_audit = match_scale_to_probability(
        scale_master=scale_master,
        probability_data=probability_data,
    )

    # 仅保留“成功匹配”且“六个量表全部完成”的被试用于 Spearman 相关分析。
    # 这样所有“亚型概率 × 量表总分”检验都基于同一批完整样本，
    # 避免不同量表因缺失情况不同而使用不同的被试集合。
    matched_mask = merged["person_key"].notna()
    complete_scale_mask = merged[SCALE_COLUMNS].notna().all(axis=1)
    analysis_mask = matched_mask & complete_scale_mask
    analysis_data = merged[analysis_mask].copy()

    print(f"成功匹配被试数：{int(matched_mask.sum())}")
    print(f"完成全部 6 个量表的匹配被试数：{int(analysis_mask.sum())}")

    if analysis_data.empty:
        raise ValueError(
            "没有任何被试同时满足“成功匹配”且“完成全部 6 个量表”。"
        )

    # 4. Spearman 相关分析
    spearman_results = run_spearman_probability_analysis(
        data=analysis_data,
        probability_columns=probability_columns,
        scale_columns=SCALE_COLUMNS,
    )

    # 5. 汇总并导出
    tables = build_output_tables(
        scale_master=scale_master,
        probability_data=probability_data,
        analysis_data=analysis_data,
        match_audit=match_audit,
        probability_columns=probability_columns,
        spearman_results=spearman_results,
    )
    export_tables(OUTPUT_XLSX, tables)

    nominal_n = len(tables["05_nominal_p_lt_05"])
    fdr_n = len(tables["06_fdr_significant"])

    print("\n分析完成。")
    print(f"输出文件：{OUTPUT_XLSX}")
    print(f"成功匹配被试数：{int(match_audit['matched'].sum())}")
    print(f"Spearman 检验数量：{len(spearman_results)}")
    print(f"未校正 p < .05 的结果数量：{nominal_n}")
    print(f"FDR q < .05 的结果数量：{fdr_n}")


if __name__ == "__main__":
    main()
