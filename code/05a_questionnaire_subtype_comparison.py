# -*- coding: utf-8 -*-
"""
不同人群亚型在六个量表总分上的差异分析。

统计流程保持不变：
- 读取六个量表原始总分和人群亚型分类；
- 按标准化姓名匹配；
- 仅对完成全部六个量表且成功匹配亚型的被试进行统计分析；
- 每个量表执行 Welch ANOVA；
- 满足当前 POSTHOC_TRIGGER_P 条件时执行两两 Welch t 检验；
- 输出 eta squared、Hedges' g 和 Benjamini-Hochberg FDR 结果；
- 汇总导出为一个 Excel 工作簿。

原始 Excel schema 保持中文，脚本不会修改源文件或量表数据结构。
"""

from __future__ import annotations

import math
import re
import warnings
from functools import reduce
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from scipy import stats

warnings.filterwarnings("ignore", category=RuntimeWarning)


# =============================================================================
# 1. 路径设置：运行前主要修改这里
# =============================================================================
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = (SCRIPT_DIR / ".." / "data").resolve()

SCALE_DIR = DATA_DIR

CLASSIFICATION_FILE = (
    SCRIPT_DIR
    / "Chapter4_Output"
    / "Table_S4_subject_classification_and_features.csv"
)

OUTPUT_XLSX = (
    SCRIPT_DIR
    / "Chapter5_Output"
    / "subtype_scale_welch_anova_results.xlsx"
)

# 如果你想按 raw_cluster 分组，可以改成 "raw_cluster"
CATEGORY_COLUMN = "subtype_cn"

# 每组至少多少人，才纳入 Welch ANOVA / Welch t 检验
MIN_GROUP_N = 2

# 总体 Welch ANOVA 的原始 p 值小于该阈值时，才做两两 Welch t 检验
POSTHOC_TRIGGER_P = 1


# =============================================================================
# 2. 六个量表文件名与量表方向
# =============================================================================

SCALE_FILES = {
    "AttentionControl": "01_16条目注意控制量表_匿名化.xlsx",
    "MentalStrength": "02_心力自评量表_匿名化.xlsx",
    "Depression": "03_抑郁情绪自评量表_匿名化.xlsx",
    "Anxiety": "04_焦虑情绪自评量表_匿名化.xlsx",
    "Insomnia": "05_失眠严重程度指数自评量表_匿名化.xlsx",
    "MMSE": "06_简易精神状态评价量表_匿名化.xlsx",
}

SCALE_INFO = {
    "AttentionControl": {
        "scale_cn": "注意控制",
        "score_direction": "higher_is_better",
    },
    "MentalStrength": {
        "scale_cn": "心力",
        "score_direction": "higher_is_better",
    },
    "Depression": {
        "scale_cn": "抑郁",
        "score_direction": "lower_is_better",
    },
    "Anxiety": {
        "scale_cn": "焦虑",
        "score_direction": "lower_is_better",
    },
    "Insomnia": {
        "scale_cn": "失眠",
        "score_direction": "lower_is_better",
    },
    "MMSE": {
        "scale_cn": "MMSE",
        "score_direction": "higher_is_better",
    },
}

SCALE_COLUMNS = list(SCALE_FILES.keys())

NAME_CANDIDATES = ["姓名", "名字", "name", "Name"]
PHONE_CANDIDATES = ["手机号", "手机号码", "电话", "phone", "Phone"]
SCORE_CANDIDATES = ["总分", "总得分", "合计", "score", "total", "Total"]
DATE_CANDIDATES = ["填写日期", "日期", "date", "Date"]
SUBJECT_CANDIDATES = ["subject", "姓名", "被试", "被试姓名", "Subject"]

CLASSIFICATION_EXTRA_COLUMNS = [
    "raw_cluster",
    "subtype_en",
    "subtype_cn",
    "screen_group",
    "PC1",
    "PC2",
    "PC3",
    "PC4",
]


# =============================================================================
# 3. 通用工具函数
# =============================================================================

def clean_column_name(value: Any) -> str:
    """清理列名。"""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_name(value: Any) -> str:
    """
    标准化姓名，用于匹配。
    会去掉空格、中文间隔点等。
    """
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"[\s·•・　]+", "", text)
    return text


def normalize_phone(value: Any) -> str:
    """标准化手机号。这里主要用于量表内部去重。"""
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D", "", text)

    if len(digits) > 11:
        digits = digits[-11:]

    return digits


def read_csv_flexible(path: Path) -> pd.DataFrame:
    """兼容 utf-8-sig、utf-8、gb18030、gbk 读取 CSV。"""
    last_error = None

    for encoding in ["utf-8-sig", "utf-8", "gb18030", "gbk"]:
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"无法读取 CSV 文件：{path}") from last_error


def first_existing_column(
    df: pd.DataFrame,
    candidates: Iterable[str],
) -> str | None:
    """从候选列名中找到第一个真实存在的列。"""
    lookup = {
        clean_column_name(col).lower(): col
        for col in df.columns
    }

    for candidate in candidates:
        key = clean_column_name(candidate).lower()
        if key in lookup:
            return lookup[key]

    return None


def safe_numeric(series: pd.Series) -> pd.Series:
    """转成数值型，无法转换的设为 NaN。"""
    return pd.to_numeric(series, errors="coerce").replace(
        [np.inf, -np.inf],
        np.nan,
    )


def parse_excel_or_text_date(series: pd.Series) -> pd.Series:
    """
    兼容 Excel 序列日期和普通文本日期。
    量表里的“填写日期”常见形式是 Excel 序列号，例如 46013。
    """
    numeric = pd.to_numeric(series, errors="coerce")

    excel_date = pd.to_datetime(
        numeric,
        unit="D",
        origin="1899-12-30",
        errors="coerce",
    )

    text_part = series.where(numeric.isna())
    text_date = pd.to_datetime(text_part, errors="coerce")

    return excel_date.combine_first(text_date)


def benjamini_hochberg(p_values: Iterable[float]) -> np.ndarray:
    """
    Benjamini-Hochberg FDR 校正。
    这里不是用来决定是否做 post-hoc，post-hoc 仍按总体原始 p < POSTHOC_TRIGGER_P 触发。
    q 值只是作为参考输出。
    """
    p = np.asarray(list(p_values), dtype=float)
    q = np.full_like(p, np.nan)

    valid = np.isfinite(p)
    if valid.sum() == 0:
        return q

    valid_p = p[valid]
    order = np.argsort(valid_p)
    ranked = valid_p[order]

    adjusted = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    adjusted = np.clip(adjusted, 0, 1)

    restored = np.empty_like(adjusted)
    restored[order] = adjusted
    q[valid] = restored

    return q


# =============================================================================
# 4. 统计函数：Welch ANOVA、Welch t、效应量
# =============================================================================

def welch_anova(groups: list[np.ndarray]) -> tuple[float, float, float, float]:
    """
    手动实现 Welch ANOVA。

    返回：
    F, df1, df2, p
    """
    clean_groups = [
        np.asarray(group, dtype=float)
        for group in groups
    ]

    clean_groups = [
        group[np.isfinite(group)]
        for group in clean_groups
        if np.isfinite(group).sum() >= MIN_GROUP_N
    ]

    k = len(clean_groups)

    if k < 2:
        return np.nan, np.nan, np.nan, np.nan

    n = np.array([len(group) for group in clean_groups], dtype=float)
    means = np.array([np.mean(group) for group in clean_groups], dtype=float)
    variances = np.array([np.var(group, ddof=1) for group in clean_groups], dtype=float)

    # 避免某组方差为 0 导致除零
    variances = np.where(variances <= 1e-12, 1e-12, variances)

    weights = n / variances
    weight_sum = np.sum(weights)
    weighted_mean = np.sum(weights * means) / weight_sum

    numerator = np.sum(weights * (means - weighted_mean) ** 2) / (k - 1)

    correction_sum = np.sum(
        ((1 - weights / weight_sum) ** 2) / (n - 1)
    )

    denominator = 1 + (
        2 * (k - 2) / (k**2 - 1)
    ) * correction_sum

    F = numerator / denominator
    df1 = k - 1
    df2 = (k**2 - 1) / (3 * correction_sum) if correction_sum > 0 else np.inf

    p = stats.f.sf(F, df1, df2)

    return float(F), float(df1), float(df2), float(p)


def eta_squared_from_groups(groups: list[np.ndarray]) -> float:
    """
    计算 eta squared / η²。

    η² = SS_between / SS_total
    """
    clean_groups = [
        np.asarray(group, dtype=float)
        for group in groups
    ]

    clean_groups = [
        group[np.isfinite(group)]
        for group in clean_groups
        if np.isfinite(group).sum() >= 1
    ]

    if len(clean_groups) < 2:
        return np.nan

    all_values = np.concatenate(clean_groups)

    if len(all_values) < 2:
        return np.nan

    grand_mean = np.mean(all_values)
    ss_total = np.sum((all_values - grand_mean) ** 2)

    if ss_total <= 0:
        return np.nan

    ss_between = 0.0

    for group in clean_groups:
        ss_between += len(group) * (np.mean(group) - grand_mean) ** 2

    return float(ss_between / ss_total)


def welch_t_test(x: np.ndarray, y: np.ndarray) -> tuple[float, float, float]:
    """
    两独立样本 Welch t 检验。

    返回：
    t, df, p
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    x = x[np.isfinite(x)]
    y = y[np.isfinite(y)]

    if len(x) < MIN_GROUP_N or len(y) < MIN_GROUP_N:
        return np.nan, np.nan, np.nan

    mean_x = np.mean(x)
    mean_y = np.mean(y)

    var_x = np.var(x, ddof=1)
    var_y = np.var(y, ddof=1)

    if var_x <= 1e-12:
        var_x = 1e-12
    if var_y <= 1e-12:
        var_y = 1e-12

    se = math.sqrt(var_x / len(x) + var_y / len(y))

    if se <= 0:
        return np.nan, np.nan, np.nan

    t_value = (mean_x - mean_y) / se

    numerator = (var_x / len(x) + var_y / len(y)) ** 2
    denominator = (
        (var_x / len(x)) ** 2 / (len(x) - 1)
        + (var_y / len(y)) ** 2 / (len(y) - 1)
    )

    df = numerator / denominator if denominator > 0 else np.nan
    p = 2 * stats.t.sf(abs(t_value), df)

    return float(t_value), float(df), float(p)


def hedges_g(x: np.ndarray, y: np.ndarray) -> float:
    """
    计算 Hedges' g。

    g > 0 表示 group_a 均值高于 group_b。
    g < 0 表示 group_a 均值低于 group_b。
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    x = x[np.isfinite(x)]
    y = y[np.isfinite(y)]

    n_x = len(x)
    n_y = len(y)

    if n_x < 2 or n_y < 2:
        return np.nan

    var_x = np.var(x, ddof=1)
    var_y = np.var(y, ddof=1)

    pooled_df = n_x + n_y - 2

    if pooled_df <= 0:
        return np.nan

    pooled_variance = (
        (n_x - 1) * var_x
        + (n_y - 1) * var_y
    ) / pooled_df

    if pooled_variance <= 0:
        return np.nan

    cohen_d = (np.mean(x) - np.mean(y)) / math.sqrt(pooled_variance)

    # 小样本校正
    correction = 1 - 3 / (4 * (n_x + n_y) - 9)

    return float(cohen_d * correction)


def interpret_pairwise_difference(
    scale: str,
    group_a: str,
    group_b: str,
    mean_diff: float,
) -> str:
    """
    根据量表方向解释均值差。

    mean_diff = group_a - group_b
    """
    if not np.isfinite(mean_diff):
        return ""

    direction = SCALE_INFO[scale]["score_direction"]

    if abs(mean_diff) < 1e-12:
        return f"{group_a} 与 {group_b} 均值几乎相同"

    if direction == "higher_is_better":
        if mean_diff > 0:
            return f"{group_a} 高于 {group_b}，表示 {group_a} 在该量表上状态更好"
        else:
            return f"{group_a} 低于 {group_b}，表示 {group_a} 在该量表上状态更差"

    if direction == "lower_is_better":
        if mean_diff > 0:
            return f"{group_a} 高于 {group_b}，表示 {group_a} 症状更重 / 状态更差"
        else:
            return f"{group_a} 低于 {group_b}，表示 {group_a} 症状更轻 / 状态更好"

    return ""


# =============================================================================
# 5. 读取六个量表
# =============================================================================

def load_one_scale(scale: str, file_path: Path) -> pd.DataFrame:
    """
    读取一个量表文件，返回：
    name_norm, name_raw, phone_norm, total_score, scale
    """
    if not file_path.exists():
        raise FileNotFoundError(f"量表文件不存在：{file_path}")

    sheets = pd.read_excel(
        file_path,
        sheet_name=None,
        dtype=object,
        engine="openpyxl",
    )

    all_parts = []

    for sheet_name, raw in sheets.items():
        df = raw.copy()
        df.columns = [clean_column_name(col) for col in df.columns]

        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

        name_col = first_existing_column(df, NAME_CANDIDATES)
        phone_col = first_existing_column(df, PHONE_CANDIDATES)
        score_col = first_existing_column(df, SCORE_CANDIDATES)
        date_col = first_existing_column(df, DATE_CANDIDATES)

        if name_col is None or score_col is None:
            continue

        part = pd.DataFrame({
            "scale": scale,
            "source_file": file_path.name,
            "source_sheet": sheet_name,
            "name_raw": df[name_col],
            "name_norm": df[name_col].map(normalize_name),
            "phone_norm": df[phone_col].map(normalize_phone) if phone_col else "",
            "total_score": safe_numeric(df[score_col]),
        })

        if date_col:
            part["fill_date"] = df[date_col]
            part["fill_date_parsed"] = parse_excel_or_text_date(df[date_col])
        else:
            part["fill_date"] = np.nan
            part["fill_date_parsed"] = pd.NaT

        part = part[
            (part["name_norm"] != "")
            & part["total_score"].notna()
        ].copy()

        all_parts.append(part)

    if not all_parts:
        raise ValueError(
            f"在量表文件中没有找到有效的姓名列和总分列：{file_path}"
        )

    scale_df = pd.concat(all_parts, ignore_index=True)

    # 同一姓名重复填写时，保留填写日期最新的一条；
    # 如果没有日期，则保留文件中最后出现的一条。
    scale_df["_original_order"] = np.arange(len(scale_df))

    scale_df = scale_df.sort_values(
        ["name_norm", "fill_date_parsed", "_original_order"],
        na_position="first",
    )

    duplicate_count = scale_df.duplicated("name_norm").sum()
    if duplicate_count > 0:
        print(
            f"[提示] {SCALE_INFO[scale]['scale_cn']} 中发现 "
            f"{duplicate_count} 条重复姓名记录，已按最新填写日期保留。"
        )

    scale_df = scale_df.drop_duplicates("name_norm", keep="last")

    scale_df = scale_df[[
        "name_norm",
        "name_raw",
        "phone_norm",
        "total_score",
        "scale",
        "source_file",
        "source_sheet",
        "fill_date",
        "fill_date_parsed",
    ]].copy()

    return scale_df


def load_all_scales() -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    读取六个量表，返回：
    1. wide_df：每个被试一行，六个量表总分为六列
    2. long_df：长表，每个被试-量表一行
    """
    long_tables = []

    for scale, filename in SCALE_FILES.items():
        file_path = SCALE_DIR / filename

        print(f"正在读取量表：{SCALE_INFO[scale]['scale_cn']} -> {file_path}")

        one = load_one_scale(scale, file_path)
        long_tables.append(one)

    long_df = pd.concat(long_tables, ignore_index=True)

    score_tables = []

    for scale in SCALE_COLUMNS:
        temp = long_df[long_df["scale"] == scale][[
            "name_norm",
            "total_score",
        ]].copy()

        temp = temp.rename(columns={"total_score": scale})
        score_tables.append(temp)

    wide_df = reduce(
        lambda left, right: pd.merge(left, right, on="name_norm", how="outer"),
        score_tables,
    )

    # 保留一个原始姓名，便于检查
    identity = (
        long_df.sort_values(["name_norm", "scale"])
        .groupby("name_norm", as_index=False)
        .agg(
            name_raw=("name_raw", "first"),
            phone_norm=("phone_norm", lambda x: next((v for v in x if v), "")),
        )
    )

    wide_df = identity.merge(wide_df, on="name_norm", how="right")

    wide_df["n_scales_available"] = wide_df[SCALE_COLUMNS].notna().sum(axis=1)

    return wide_df, long_df


# =============================================================================
# 6. 读取人群亚型分类文件
# =============================================================================

def load_classification(path: Path) -> pd.DataFrame:
    """
    读取 Table_S3_4_subject_classification_and_features.csv。

    默认使用：
    - subject 作为姓名
    - subtype_cn 作为亚型类别
    """
    if not path.exists():
        raise FileNotFoundError(f"分类文件不存在：{path}")

    df = read_csv_flexible(path)
    df.columns = [clean_column_name(col) for col in df.columns]

    subject_col = first_existing_column(
        df,
        SUBJECT_CANDIDATES,
    )

    if subject_col is None:
        raise ValueError("分类文件中没有找到 subject / 姓名 列。")

    if CATEGORY_COLUMN not in df.columns:
        raise ValueError(
            f"分类文件中没有找到指定分组列：{CATEGORY_COLUMN}\n"
            f"当前可用列名为：{list(df.columns)}"
        )

    output = df.copy()

    output["subject_raw"] = output[subject_col]
    output["name_norm"] = output[subject_col].map(normalize_name)

    output["subtype"] = output[CATEGORY_COLUMN].astype(str).str.strip()
    output.loc[output["subtype"].isin(["", "nan", "None"]), "subtype"] = np.nan

    output = output[
        (output["name_norm"] != "")
        & output["subtype"].notna()
    ].copy()

    duplicate_count = output.duplicated("name_norm").sum()
    if duplicate_count > 0:
        print(
            f"[警告] 分类文件中发现 {duplicate_count} 条重复 subject。"
            "代码将保留第一次出现的记录，请你后续检查。"
        )
        output = output.drop_duplicates("name_norm", keep="first")

    keep_cols = [
        "name_norm",
        "subject_raw",
        "subtype",
    ]

    extra_cols = [
        col
        for col in CLASSIFICATION_EXTRA_COLUMNS
        if col in output.columns
    ]

    keep_cols += [
        col for col in extra_cols
        if col not in keep_cols
    ]

    return output[keep_cols].copy()


# =============================================================================
# 7. 主分析函数
# =============================================================================

def _scale_descriptives(
    subset: pd.DataFrame,
    scale: str,
    subtype_labels: list[str],
) -> tuple[list[dict[str, Any]], list[str], list[np.ndarray]]:
    """生成单个量表的分组描述统计和可用于检验的有效组。"""
    scale_cn = SCALE_INFO[scale]["scale_cn"]
    score_direction = SCALE_INFO[scale]["score_direction"]

    rows = []
    valid_labels = []
    valid_groups = []

    for subtype in subtype_labels:
        values = safe_numeric(
            subset.loc[subset["subtype"] == subtype, scale]
        ).dropna().to_numpy(dtype=float)

        rows.append({
            "scale": scale,
            "scale_cn": scale_cn,
            "score_direction": score_direction,
            "subtype": subtype,
            "n": len(values),
            "mean": float(np.mean(values)) if len(values) > 0 else np.nan,
            "sd": float(np.std(values, ddof=1)) if len(values) > 1 else np.nan,
            "median": float(np.median(values)) if len(values) > 0 else np.nan,
            "min": float(np.min(values)) if len(values) > 0 else np.nan,
            "max": float(np.max(values)) if len(values) > 0 else np.nan,
        })

        if len(values) >= MIN_GROUP_N:
            valid_labels.append(subtype)
            valid_groups.append(values)

    return rows, valid_labels, valid_groups


def _pairwise_posthoc(
    subset: pd.DataFrame,
    scale: str,
    valid_labels: list[str],
) -> list[dict[str, Any]]:
    """对有效亚型组合执行两两 Welch t 检验。"""
    rows = []
    scale_cn = SCALE_INFO[scale]["scale_cn"]
    score_direction = SCALE_INFO[scale]["score_direction"]

    for i, subtype_a in enumerate(valid_labels):
        for subtype_b in valid_labels[i + 1:]:
            x = safe_numeric(
                subset.loc[subset["subtype"] == subtype_a, scale]
            ).dropna().to_numpy(dtype=float)
            y = safe_numeric(
                subset.loc[subset["subtype"] == subtype_b, scale]
            ).dropna().to_numpy(dtype=float)

            t_value, welch_df, pair_p = welch_t_test(x, y)
            mean_a = float(np.mean(x)) if len(x) > 0 else np.nan
            mean_b = float(np.mean(y)) if len(y) > 0 else np.nan
            mean_diff = mean_a - mean_b

            rows.append({
                "scale": scale,
                "scale_cn": scale_cn,
                "score_direction": score_direction,
                "subtype_a": subtype_a,
                "subtype_b": subtype_b,
                "n_a": len(x),
                "n_b": len(y),
                "mean_a": mean_a,
                "mean_b": mean_b,
                "mean_difference_a_minus_b": mean_diff,
                "Welch_t": t_value,
                "df": welch_df,
                "p": pair_p,
                "Hedges_g": hedges_g(x, y),
                "interpretation": interpret_pairwise_difference(
                    scale=scale,
                    group_a=subtype_a,
                    group_b=subtype_b,
                    mean_diff=mean_diff,
                ),
            })

    return rows


def analyze_scale_by_subtype(
    data: pd.DataFrame,
    scale: str,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """对单个量表执行描述统计、Welch ANOVA 和条件式 post-hoc。"""
    subset = data[["subtype", scale]].copy()
    subset[scale] = safe_numeric(subset[scale])
    subset = subset.dropna(subset=["subtype", scale])

    subtype_labels = sorted(subset["subtype"].unique().tolist())
    descriptives, valid_labels, valid_groups = _scale_descriptives(
        subset,
        scale,
        subtype_labels,
    )

    F, df1, df2, p = welch_anova(valid_groups)
    eta2 = eta_squared_from_groups(valid_groups)

    overall = {
        "scale": scale,
        "scale_cn": SCALE_INFO[scale]["scale_cn"],
        "score_direction": SCALE_INFO[scale]["score_direction"],
        "n_total": int(subset[scale].notna().sum()),
        "n_subtypes_total": int(len(subtype_labels)),
        "n_subtypes_used": int(len(valid_groups)),
        "Welch_F": F,
        "df1": df1,
        "df2": df2,
        "p": p,
        "eta_squared": eta2,
        "posthoc_performed": bool(np.isfinite(p) and p < POSTHOC_TRIGGER_P),
    }

    posthoc_rows = (
        _pairwise_posthoc(subset, scale, valid_labels)
        if np.isfinite(p) and p < POSTHOC_TRIGGER_P
        else []
    )

    return overall, posthoc_rows, descriptives

def run_all_analyses(merged_data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """对六个量表循环执行分析。"""
    overall_rows = []
    posthoc_rows = []
    descriptive_rows = []

    for scale in SCALE_COLUMNS:
        overall, posthoc, descriptives = analyze_scale_by_subtype(
            merged_data,
            scale,
        )

        overall_rows.append(overall)
        posthoc_rows.extend(posthoc)
        descriptive_rows.extend(descriptives)

    overall_df = pd.DataFrame(overall_rows)
    posthoc_df = pd.DataFrame(posthoc_rows)
    descriptives_df = pd.DataFrame(descriptive_rows)

    if not overall_df.empty:
        overall_df["q_FDR_across_6_scales"] = benjamini_hochberg(overall_df["p"])
        overall_df["p_lt_05"] = overall_df["p"] < 0.05
        overall_df["q_FDR_lt_05"] = overall_df["q_FDR_across_6_scales"] < 0.05

    if not posthoc_df.empty:
        posthoc_df["q_FDR_across_all_posthoc"] = benjamini_hochberg(posthoc_df["p"])
        posthoc_df["p_lt_05"] = posthoc_df["p"] < 0.05
        posthoc_df["q_FDR_lt_05"] = posthoc_df["q_FDR_across_all_posthoc"] < 0.05

    return overall_df, posthoc_df, descriptives_df


# =============================================================================
# 8. 导出 Excel
# =============================================================================

def autosize_excel_columns(path: Path) -> None:
    """简单美化 Excel 列宽和表头。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in ws.columns:
            values = [
                str(cell.value) if cell.value is not None else ""
                for cell in column_cells[:200]
            ]
            max_len = max([len(v) for v in values], default=10)
            width = min(max(max_len + 2, 10), 40)
            ws.column_dimensions[column_cells[0].column_letter].width = width

    wb.save(path)


def make_output_metadata(
    merged_data: pd.DataFrame,
    classification: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """构建工作簿中的分析说明、量表规则和样本量表。"""
    readme = pd.DataFrame({
        "item": [
            "analysis_goal",
            "group_column",
            "matching_rule",
            "scale_score",
            "overall_test",
            "posthoc_rule",
            "posthoc_test",
            "overall_effect_size",
            "posthoc_effect_size",
            "score_direction",
        ],
        "value": [
            "Compare six questionnaire total scores across population subtypes",
            CATEGORY_COLUMN,
            "Questionnaire 姓名 matched to classification subject after name normalization",
            "Raw total score, no standardization and no reverse transformation",
            "Welch ANOVA",
            f"Pairwise Welch t tests are performed only when overall raw p < {POSTHOC_TRIGGER_P}",
            "Welch independent-samples t test",
            "eta squared = SS_between / SS_total",
            "Hedges' g",
            "AttentionControl/MentalStrength/MMSE: higher is better; "
            "Depression/Anxiety/Insomnia: lower is better",
        ],
    })

    scale_rules = pd.DataFrame([
        {
            "scale": scale,
            "scale_cn": SCALE_INFO[scale]["scale_cn"],
            "score_direction": SCALE_INFO[scale]["score_direction"],
        }
        for scale in SCALE_COLUMNS
    ])

    sample_counts = pd.DataFrame([
        {"item": "classification_subjects", "value": len(classification)},
        {
            "item": "scale_subjects_total",
            "value": merged_data["name_norm"].nunique(),
        },
        {
            "item": "matched_subjects_with_subtype",
            "value": int(merged_data["subtype"].notna().sum()),
        },
        {
            "item": "subjects_with_all_6_scales_and_subtype",
            "value": int(
                merged_data[["subtype"] + SCALE_COLUMNS]
                .dropna()
                .shape[0]
            ),
        },
    ])

    return readme, scale_rules, sample_counts


def export_results(
    output_path: Path,
    merged_data: pd.DataFrame,
    scale_long: pd.DataFrame,
    classification: pd.DataFrame,
    descriptives: pd.DataFrame,
    overall: pd.DataFrame,
    posthoc: pd.DataFrame,
) -> None:
    """导出所有结果到单个 Excel 工作簿。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    readme, scale_rules, sample_counts = make_output_metadata(
        merged_data,
        classification,
    )
    unmatched_scale_subjects = merged_data[
        merged_data["subtype"].isna()
    ].copy()

    sheets = {
        "00_README": readme,
        "01_scale_rules": scale_rules,
        "02_sample_counts": sample_counts,
        "03_matched_data": merged_data,
        "04_descriptives": descriptives,
        "05_welch_anova": overall,
        "06_posthoc_welch_t": posthoc,
        "07_unmatched_scale_subjects": unmatched_scale_subjects,
        "08_scale_long_raw": scale_long,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheets.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    autosize_excel_columns(output_path)


# =============================================================================
# 9. 主程序
# =============================================================================

def main() -> None:
    print("=" * 80)
    print("开始分析：不同人群亚型在六个量表总分上的差异")
    print("=" * 80)

    print("\n[1/5] 读取六个量表文件")
    scale_wide, scale_long = load_all_scales()

    print("\n[2/5] 读取人群亚型分类文件")
    classification = load_classification(CLASSIFICATION_FILE)

    print("\n[3/5] 按姓名匹配量表数据与人群亚型")
    merged = scale_wide.merge(
        classification,
        on="name_norm",
        how="left",
    )

    matched_n = int(merged["subtype"].notna().sum())
    total_n = len(merged)

    print(f"量表被试数：{total_n}")
    print(f"成功匹配到亚型的被试数：{matched_n}")
    print(f"未匹配到亚型的被试数：{total_n - matched_n}")

    # 仅保留“六个量表全部完成”且“成功匹配到亚型”的被试用于统计分析
    complete_scale_mask = merged[SCALE_COLUMNS].notna().all(axis=1)
    complete_scale_n = int(complete_scale_mask.sum())
    complete_and_matched_mask = complete_scale_mask & merged["subtype"].notna()
    complete_and_matched_n = int(complete_and_matched_mask.sum())

    print(f"完成全部 6 个量表的被试数：{complete_scale_n}")
    print(f"完成全部 6 个量表且成功匹配亚型的被试数：{complete_and_matched_n}")

    analysis_data = merged[complete_and_matched_mask].copy()

    if analysis_data.empty:
        raise ValueError(
            "没有任何被试同时满足“完成全部 6 个量表”且“成功匹配到人群亚型”。"
        )

    print("\n[4/5] 执行 Welch ANOVA 和必要的两两 Welch t 检验")
    overall_df, posthoc_df, descriptives_df = run_all_analyses(analysis_data)

    print("\nWelch ANOVA 总体结果：")
    display_cols = [
        "scale_cn",
        "n_total",
        "n_subtypes_used",
        "Welch_F",
        "df1",
        "df2",
        "p",
        "eta_squared",
        "posthoc_performed",
    ]
    print(overall_df[display_cols].to_string(index=False))

    print("\n[5/5] 导出 Excel 结果")
    export_results(
        output_path=OUTPUT_XLSX,
        merged_data=merged,
        scale_long=scale_long,
        classification=classification,
        descriptives=descriptives_df,
        overall=overall_df,
        posthoc=posthoc_df,
    )

    print("\n分析完成。")
    print(f"结果文件已保存到：{OUTPUT_XLSX}")

    if posthoc_df.empty:
        print(f"\n没有任何量表满足总体原始 p < {POSTHOC_TRIGGER_P}，因此未执行两两 Welch t 检验。")
    else:
        print(f"\n已生成两两 Welch t 检验结果，共 {len(posthoc_df)} 条比较。")


if __name__ == "__main__":
    main()
